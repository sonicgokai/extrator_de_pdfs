import pdfplumber as pb
import openpyxl as x
import os

# buscar os arquivos  pdf
folder = os.listdir("pdfs")
# Varrendo a pasta pdfs no computador e  verificando a  extensao dos arquivos
# Vse sao realmente PDFs
for file in folder:
    if file.lower().endswith(".pdf"):
        try:
            # Abrindo o arquivo Excel
            excel = x.load_workbook("base de  dados inspecoes.xlsx")
            sheet = excel.active
            line_start = len(sheet["A"]) + 1
            # Varrendo os arquivos PDFs e extraindo os dados
            pdf = pb.open(f"pdfs\\{file}")
            page = pdf.pages[0]
            data = page.extract_table()
            # Varrendo o arquivo PDF
            for i, d in enumerate(data[1:], start=line_start):
                # Se os registros dos  PDF  nao foren nulos s
                if d[0] == "":
                    pass
                else:
                    sheet.cell(row=i, column=1).value = d[0]
                    sheet.cell(row=i, column=2).value = d[1]
                    sheet.cell(row=i, column=3).value = d[2]
                    sheet.cell(row=i, column=4).value = d[3]
                    sheet.cell(row=i, column=5).value = d[4]
            pdf.close()
            excel.save("base de  dados inspecoes.xlsx")
            excel.close()
        except Exception as e:
            with open("log.txt", "a") as log:
                log.write(f"Houve um erro para extrair os dados do arquivo {file}\n")
                log.write(f"Erro: {e}")

    else:
        with open("log.txt", "a") as log:
            log.write(f"O arquivo {file} nao eh um PDF  valido!\n")
